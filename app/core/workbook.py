"""Exact blueprint-compliant three-sheet workbook builder.

Adapted from blueprint v1.0 historical_model_builder.py. It preserves the
canonical-key compiler and adds year-specific formula support, safe source text,
and validation markers without adding sheets to the main workbook.
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Any,Mapping
from openpyxl import Workbook,load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.utils import get_column_letter
from .security import escape_excel_text
from .resources import load_json

SHEET_ORDER=["INCOME STATEMENT","BALANCE SHEET","CASH FLOW STATEMENT"]
STYLE=load_json("style_config.json"); CONTRACT=load_json("output_contract.json")
CURRENCY_FMT=CONTRACT["style"]["currency_format"]; PERCENT_FMT=CONTRACT["style"]["percent_format"]; COUNT_FMT=CONTRACT["style"]["count_format"]
ACTUAL_FILL=CONTRACT["style"]["actual_fill"].lstrip("#"); SUPPLEMENT_FILL=CONTRACT["style"]["supplement_fill"].lstrip("#")
BLACK="000000"; WHITE="FFFFFF"; THIN=Side(style="thin",color=BLACK); MEDIUM=Side(style="medium",color=BLACK)
ROW_TYPES={"section_header","subgroup_header","component","subtotal","major_total","supplement_header","metric","error_check"}
VALUE_TYPES={"currency","percent","count","text","none"}
@dataclass
class BuildContext: years:list[str]; row_map:dict[str,dict[str,int]]

def validate_payload(payload:Mapping[str,Any])->None:
    years=payload.get("years")
    if not isinstance(years,list) or not years or len(set(years))!=len(years):raise ValueError("years must be unique non-empty list")
    if set(payload.get("statements",{}))!=set(SHEET_ORDER):raise ValueError("exactly three required statement sheets must be present")
    if any("forecast" in str(y).casefold() or "estimate" in str(y).casefold() for y in years):raise ValueError("forecast periods are prohibited")
    for sheet in SHEET_ORDER:
        seen=set()
        for row in payload["statements"][sheet]:
            key=row.get("key")
            if not key or key in seen:raise ValueError(f"invalid/duplicate key {sheet}:{key}")
            seen.add(key)
            if row.get("row_type","component") not in ROW_TYPES:raise ValueError(f"unsupported row type {row.get('row_type')}")
            if row.get("value_type","currency") not in VALUE_TYPES:raise ValueError(f"unsupported value type {row.get('value_type')}")

def included(row:Mapping[str,Any],years:list[str])->bool:
    if row.get("include") is False:return False
    if row.get("row_type") in {"section_header","subgroup_header","supplement_header"}:return bool(row.get("has_included_children"))
    if row.get("formula") or row.get("formulas"):return True
    vals=row.get("values",{});return any(y in vals and vals[y] is not None for y in years) or bool(row.get("required_blank_structure"))

def build_row_map(payload):
    out={};years=payload["years"]
    for sheet in SHEET_ORDER:
        out[sheet]={};r=3
        for row in payload["statements"][sheet]:
            if not included(row,years):continue
            out[sheet][row["key"]]=r;r+=1+(1 if row.get("blank_after") else 0)
    return out

def _sheet_ref(name):return "'"+name.replace("'","''")+"'"
def compile_formula(defn:Mapping[str,Any],sheet:str,year_index:int,ctx:BuildContext)->str:
    typ=defn.get("type")
    def ref(key,target_sheet=None,target_year=None):
        s=target_sheet or sheet;i=year_index if target_year is None else target_year
        if key not in ctx.row_map[s]:raise KeyError(f"formula references missing key {s}.{key}")
        return f"{_sheet_ref(s)}!{get_column_letter(3+i)}{ctx.row_map[s][key]}"
    if typ=="sum_keys":return "=SUM("+",".join(ref(k,defn.get("sheet")) for k in defn.get("keys",[]))+")"
    if typ=="subtract_keys":
        minuend=ref(defn["minuend"],defn.get("sheet")); subs=[ref(k,defn.get("sheet")) for k in defn.get("subtrahends",[])]
        return f"={minuend}" if not subs else f"={minuend}-SUM({','.join(subs)})"
    if typ=="link_key":return "="+ref(defn["key"],defn.get("sheet"))
    if typ=="previous_year_same_key":return "" if year_index==0 else "="+ref(defn["key"],defn.get("sheet"),year_index-1)
    if typ=="custom_excel":
        template=defn.get("template","")
        if not template or any(x in template.upper() for x in ("WEBSERVICE(","HYPERLINK(","DDE","CMD|")):raise ValueError("unsafe custom Excel formula")
        rendered=template.replace("{col}",get_column_letter(3+year_index))
        for token in defn.get("references",[]):rendered=rendered.replace("{"+token["placeholder"]+"}",ref(token["key"],token.get("sheet"),token.get("year_index",year_index)))
        return rendered if rendered.startswith("=") else "="+rendered
    raise ValueError(f"unsupported formula type {typ}")

def source_comment(source:Any)->Comment|None:
    if not source:return None
    records=source if isinstance(source,list) else [source];lines=[]
    for record in records:
        if isinstance(record,str):lines.append(escape_excel_text(record));continue
        if not isinstance(record,Mapping):continue
        parts=[record.get("source_file"),f"page {record.get('page')}" if record.get("page") else None,f"note {record.get('note_number')}" if record.get("note_number") else None,f"original label: {escape_excel_text(str(record.get('original_label')))}" if record.get("original_label") else None,f"status: {record.get('status')}" if record.get("status") else None]
        lines.append(" | ".join(str(p) for p in parts if p))
    return Comment("\n".join(lines),"Annual-report extraction") if lines else None

def base_style(ws,years):
    ws.sheet_view.showGridLines=False;ws.freeze_panes="C3";ws.column_dimensions["A"].width=3;ws.column_dimensions["B"].width=62
    for i in range(len(years)):ws.column_dimensions[get_column_letter(3+i)].width=16
    ws["B2"]="YEAR";ws["B2"].font=Font(name="Arial",size=10,bold=True);ws["B2"].alignment=Alignment(horizontal="left")
    for i,year in enumerate(years):
        c1=ws.cell(1,3+i,"ACTUAL");c2=ws.cell(2,3+i,year)
        c1.fill=PatternFill("solid",fgColor=ACTUAL_FILL);c1.font=Font(name="Arial",size=11);c1.alignment=Alignment(horizontal="right");c1.border=Border(top=THIN,left=THIN)
        c2.font=Font(name="Arial",size=11,bold=True);c2.alignment=Alignment(horizontal="right")

def row_style(ws,r,row,year_count):
    rt=row.get("row_type","component");vt=row.get("value_type","currency");level=int(row.get("hierarchy_level",0));label=ws.cell(r,2)
    label.font=Font(name="Arial",size=10);label.alignment=Alignment(horizontal="left",vertical="center",indent=max(0,level));label.fill=PatternFill("solid",fgColor=WHITE)
    for col in range(3,3+year_count):
        c=ws.cell(r,col);c.font=Font(name="Arial",size=10,italic=(vt=="percent"));c.alignment=Alignment(horizontal="right",vertical="center");c.fill=PatternFill("solid",fgColor=WHITE)
        c.number_format={"currency":CURRENCY_FMT,"percent":PERCENT_FMT,"count":COUNT_FMT}.get(vt,"General")
    cells=[ws.cell(r,c) for c in range(2,3+year_count)]
    if rt=="section_header":
        for c in cells:c.font=Font(name="Arial",size=10,bold=True)
    elif rt=="subgroup_header":
        for c in cells:c.font=Font(name="Arial",size=10,bold=True);c.border=Border(top=THIN)
    elif rt in {"subtotal","error_check"}:
        for c in cells:c.font=Font(name="Arial",size=10,bold=True,italic=(vt=="percent" and c.column>=3));c.border=Border(top=THIN)
    elif rt=="major_total":
        for c in cells:c.font=Font(name="Arial",size=10,bold=True,italic=(vt=="percent" and c.column>=3));c.border=Border(top=MEDIUM)
    elif rt=="supplement_header":
        for c in cells:c.fill=PatternFill("solid",fgColor=SUPPLEMENT_FILL)
        label.font=Font(name="Arial",size=10,bold=True)

def build_workbook(payload:Mapping[str,Any],review_required:bool=False)->Workbook:
    validate_payload(payload);years=list(payload["years"]);ctx=BuildContext(years,build_row_map(payload));wb=Workbook();wb.remove(wb.active)
    for sheet in SHEET_ORDER:
        ws=wb.create_sheet(sheet);base_style(ws,years);r=3
        for row in payload["statements"][sheet]:
            if not included(row,years):continue
            ws.cell(r,2,escape_excel_text(row["label"]));rt=row.get("row_type","component");vals=row.get("values",{});formulas=row.get("formulas",{});default_formula=row.get("formula");sources=row.get("sources",{});statuses=row.get("statuses",{})
            if rt not in {"section_header","subgroup_header","supplement_header"}:
                for i,year in enumerate(years):
                    c=ws.cell(r,3+i);formula=formulas.get(year) or default_formula
                    if formula:
                        compiled=compile_formula(formula,sheet,i,ctx)
                        if compiled:c.value=compiled
                    elif year in vals and vals[year] is not None:
                        value=vals[year]
                        try:c.value=float(value) if row.get("value_type")!="text" else escape_excel_text(str(value))
                        except (TypeError,ValueError):c.value=escape_excel_text(str(value))
                    comment=source_comment(sources.get(year));
                    if comment:c.comment=comment
                    if statuses.get(year) in {"AMBIGUOUS_REVIEW_REQUIRED","MISSING_NOT_DISCLOSED"}:c.fill=PatternFill("solid",fgColor="FFF4D6")
            row_style(ws,r,row,len(years));r+=1+(1 if row.get("blank_after") else 0)
    wb.calculation.fullCalcOnLoad=True;wb.calculation.forceFullCalc=True;wb.calculation.calcMode="auto"
    wb.properties.title=("REVIEW REQUIRED - " if review_required else "")+str(payload.get("company_name",""))+" Historical Three-Statement Model"
    return wb

def assert_workbook_contract(path:Path)->list[str]:
    wb=load_workbook(path,data_only=False,read_only=False);errors=[]
    if wb.sheetnames!=SHEET_ORDER:errors.append(f"SHEETS:{wb.sheetnames}")
    for ws in wb.worksheets:
        if ws.freeze_panes!="C3":errors.append(f"FREEZE:{ws.title}:{ws.freeze_panes}")
        if ws.sheet_view.showGridLines is not False:errors.append(f"GRIDLINES:{ws.title}")
        if ws["B2"].value!="YEAR":errors.append(f"YEAR_LABEL:{ws.title}")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str) and "FORECAST" in cell.value.upper():errors.append(f"FORECAST:{ws.title}:{cell.coordinate}")
                if isinstance(cell.value,str) and cell.value.startswith("=") and any(e in cell.value for e in ("#REF!","#DIV/0!","#VALUE!","#NAME?","#N/A")):errors.append(f"FORMULA_ERROR:{ws.title}:{cell.coordinate}")
    return errors
